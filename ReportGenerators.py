import openpyxl
from sklearn.metrics import classification_report

'''
Class for generating reports based on results
'''

class ClassificationReportWriter:
    def __init__(self, y_pred, y_test, test_nouns, accuracy, output_file):
        self.y_pred = y_pred
        self.y_test = y_test
        self.test_nouns = test_nouns
        self.output_file = output_file
        self.accuracy = accuracy

    def write_report(self):
        report = classification_report(self.y_test, self.y_pred, output_dict=True, zero_division=0)

        # Get the unique class labels
        classes = sorted(set(self.y_test))

        # Create a new Excel workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Write the column headers for classification report
        worksheet['A1'] = 'Class'
        worksheet['B1'] = 'Precision'
        worksheet['C1'] = 'Recall'
        worksheet['D1'] = 'F1-score'
        worksheet['E1'] = 'Support'

        # Iterate through the classes and write the classification report to the worksheet
        row = 2
        for cls in classes:
            precision = report[cls]['precision']
            recall = report[cls]['recall']
            f1_score = report[cls]['f1-score']
            support = report[cls]['support']

            worksheet.cell(row=row, column=1, value=cls)
            worksheet.cell(row=row, column=2, value=precision)
            worksheet.cell(row=row, column=3, value=recall)
            worksheet.cell(row=row, column=4, value=f1_score)
            worksheet.cell(row=row, column=5, value=int(support))

            row += 1

        worksheet.cell(row=row,column=1,value=self.accuracy)

        # Write the column headers for incorrect predictions
        worksheet['A{}'.format(row + 1)] = 'Noun'
        worksheet['B{}'.format(row + 1)] = 'Predicted Class'
        worksheet['C{}'.format(row + 1)] = 'Correct Class'

        # Iterate through the test set and write incorrect predictions to the worksheet
        row += 2
        for noun, predicted_class, correct_class in zip(self.test_nouns, self.y_pred, self.y_test):
            if predicted_class != correct_class:
                worksheet.cell(row=row, column=1, value=noun)
                worksheet.cell(row=row, column=2, value=predicted_class)
                worksheet.cell(row=row, column=3, value=correct_class)
                row += 1

        # Save the workbook
        workbook.save(self.output_file)