import openpyxl
import gzip
import ast
import zlib

"""
DataLoader Class:
Functionality: Loads Data from .xl files and returns arrays of train and test sets
Parameters: Train File Location, Test File Location, mtime (compression parameter), compression Level (compression parameter)
Function: Load_train_data, Load_test_data, Load_train_data_compressed, Load_test_data_compressed
"""

class DataLoader:
    def __init__(self, train_file, test_file, mtime=None, compressionlevel=9):
        self.train_file = train_file
        self.test_file = test_file
        self.mtime = mtime
        self.compressionlevel = compressionlevel

    #Load training data, return nouns and classes
    def load_train_data(self):
        train_workbook = openpyxl.load_workbook(self.train_file)
        train_worksheet = train_workbook.active

        train_nouns = []
        train_classes = []

        for row_num in range(1, train_worksheet.max_row + 1):
            row_values = tuple(train_worksheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
            train_nouns.append(str(row_values[0]))
            train_classes.append(str(row_values[1]))

        return train_nouns, train_classes

    #Load test data, return test nouns and classes
    def load_test_data(self):
        test_workbook = openpyxl.load_workbook(self.test_file)
        test_worksheet = test_workbook.active

        test_nouns = []
        test_classes = []

        for row_num in range(1, test_worksheet.max_row + 1):
            row_values = tuple(test_worksheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
            test_nouns.append(str(row_values[0]))
            test_classes.append(str(row_values[1]))

        return test_nouns, test_classes
    
    #load training data compressed based on compression type, retuns compressed nouns and classes
    def load_train_data_compressed(self, compression='gzip'):
        train_workbook = openpyxl.load_workbook(self.train_file)
        train_worksheet = train_workbook.active

        train_nouns = []
        train_classes = []

        for row_num in range(1, train_worksheet.max_row + 1):
            row_values = tuple(train_worksheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
            
            match compression:
                case 'gzip':
                    train_nouns.append(str(gzip.compress(bytes(row_values[0], 'utf-8'),mtime=self.mtime,compresslevel=self.compressionlevel)))
                case 'zlib':
                    train_nouns.append(str(zlib.compress(bytes(row_values[0], 'utf-8'),level=9)))

            train_classes.append(str(row_values[1]))

        return train_nouns, train_classes
    
    #load test data compressed based on compression type, retuns compressed nouns and classes
    def load_test_data_compressed(self,compression='gzip'):
        test_workbook = openpyxl.load_workbook(self.test_file)
        test_worksheet = test_workbook.active

        test_nouns = []
        test_classes = []

        for row_num in range(1, test_worksheet.max_row + 1):
            row_values = tuple(test_worksheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]

            match compression:
                case 'gzip':
                    test_nouns.append(str(gzip.compress(bytes(row_values[0], 'utf-8'),mtime=self.mtime,compresslevel=self.compressionlevel)))
                case 'zlib':
                    test_nouns.append(str(zlib.compress(bytes(row_values[0], 'utf-8'))))
    
            test_classes.append(str(row_values[1]))

        return test_nouns, test_classes
    
    #Decompress data
    def decompress_array(compressed_array, compression):
        decompressed_array = []

        for compressed_data in compressed_array:
            try:
                # Convert the compressed data from string to bytes
                compressed_bytes = ast.literal_eval(compressed_data)
                
                match compression:
                    case 'gzip':
                        decompressed_data = gzip.decompress(compressed_bytes)
                    case 'zlib':
                        decompressed_data = zlib.decompress(compressed_bytes)
                
                # Decode the decompressed data from bytes to string
                decompressed_string = decompressed_data.decode('utf-8')
                
                decompressed_array.append(decompressed_string)
            except Exception as e:
                print(e)

        return decompressed_array
