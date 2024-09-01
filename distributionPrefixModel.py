import openpyxl
from sklearn.metrics import accuracy_score
from inputGathering import SimplePrefixKnowledgeInfused
import random

'''
A knowledge infused prefixing method, makes classification if unique otherwise makes selection based on distibution
'''

#get the language choice from input
language_choice = SimplePrefixKnowledgeInfused()

#get prefixing information and representation amount(frequency occures)
match language_choice:
    case 1:
        prefix_dictionary = [
            ('1',"mo"),
            ('1a',""),
            ('2',"ba"),
            ('2b',"bo"),
            ('3',"mo"),
            ('4',"me"),
            ('5',"le"),
            ('6',"ma"),
            ('7',"se"),
            ('8',"di"),
            ('9',"n"),
            ('9',"m"),
            ('9',""),
            ('10',"din"),
            ('10',"dim"),
            ('10',"di"),
            ('14',"bo"),
            ('16',"fa"),
            ('17',"go"),
            ('18',"mo"),
            ('n',"m"),
            ('n',"n"),
            ('n',""),
            ('24',"ga"),
        ]

        representation_dictionary = {
            '1':176,
            '1a':42,
            '2':106,
            '2b':9,
            '3':180,
            '4':110,
            '5':251,
            '6':287,
            '7':210,
            '8':105,
            '9':650,
            '10':245,
            '14':162,
            '16':4,
            '17':3,
            '18':4,
            'n':14,
            '24':2,
        }

        # Load the testing data from Excel
        test_workbook = openpyxl.load_workbook('data/80_20_Split_SingleNorthSotho/NorthernSothoSingle20%TestSet.xlsx')
        test_worksheet = test_workbook.active
    case 2:
        prefix_dictionary =  [
            ('1',"um"),
            ('1',"umu"),
            ('1a',"u"),
            ('2',"aba"),
            ('2',"abe"),
            ('2a',"o"),
            ('3',"umu"),
            ('3',"um"),
            ('4',"imi"),
            ('5',"i"),
            ('5',"ili"),
            ('6',"ama"),
            ('6',"ame"),
            ('7',"is"),
            ('7',"isi"),
            ('8',"iz"),
            ('8',"izi"),
            ('9',"in"),
            ('9',"im"),
            ('10',"izin"),
            ('10',"izim"),
            ('11',"u"),
            ('11',"ulu"),
            ('14',"ubu"),
            ('14',"ub"),
            ('15',"uku"),
            ('15',"ukw"),
        ]

        representation_dictionary = {
            '1':110,
            '1a':144,
            '2':94,
            '2a':48,
            '3':160,
            '4':87,
            '5':296,
            '6':231,
            '7':229,
            '8':167,
            '9':299,
            '10':155,
            '11':98,
            '14':60,
            '15':101,
        }

        # Load the testing data from Excel
        test_workbook = openpyxl.load_workbook('data/80_20_Split_Single/ZuluNouns20%TestSet.xlsx')
        test_worksheet = test_workbook.active


#check if prefix is unique
def is_unique(unique_prefixes, query):
    for (value,prefix) in unique_prefixes:
        if query.startswith(prefix):
            return value
    return None

#function to get out all unique prefixes
def get_unique_prefixes(prefixDict):
    uniquePrefix = []
    nonUnique = []
    for (value,prefix) in prefixDict:
        isUnique = True
        for value2, otherPrefix in prefixDict:

            if(len(otherPrefix) >= len(prefix) and (value != value2)):
                if otherPrefix.startswith(prefix):
                    isUnique = False
                    break
        if isUnique:
            uniquePrefix.append((value,prefix))
        else:
            nonUnique.append((value,prefix))

    return uniquePrefix, nonUnique


test_nouns = []
test_classes = []
predicted_classes = []

unique_prefixes, non_unique_prefixes = get_unique_prefixes(prefixDict=prefix_dictionary)

accuracy_score_total = 0


# Iterate over the rows in the testing data
for row_num in range(1, test_worksheet.max_row + 1):
    row_values = tuple(test_worksheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]

    test_classes.append(str(row_values[1]))
    test_nouns.append(str(row_values[0]))
    
    loaded_noun = str(row_values[0])
    matched_prefix = []
    matched_classes = []

    unique_prediciton = is_unique(unique_prefixes,loaded_noun)

    if unique_prediciton:
        predicted_classes.append(str(unique_prediciton))
    else:
        options = []
        for (value,key) in non_unique_prefixes:
            if loaded_noun.startswith(key):
                options.append(str(value))
        highest_rep = 0
        highest_rep_class = ''
        if options:
            for x in options:
                if(representation_dictionary[x] > highest_rep):
                    highest_rep_class = x
            predicted_classes.append(highest_rep_class)
        else:
            predicted_classes.append('na')
            
#output final accuracy
accuracy_score_total += accuracy_score(test_classes, predicted_classes)
#print(f"Accuracy: {accuracy:.4f}")

print(f"Accuracy: {accuracy_score_total:.4f}")