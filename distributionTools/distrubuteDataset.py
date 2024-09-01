
from sklearn.model_selection import train_test_split
import openpyxl

'''
code to create Representative split of data 80-20
'''


input_label = "NorthernSothoDual.xlsx"
output_label = "NorthernSothoDual"

# Load the training data from Excel
train_workbook = openpyxl.load_workbook(fr'..\data\CoreDataSets\{input_label}')
train_worksheet = train_workbook.active

nouns = []
classes = []

# Iterate over the rows in the training data
for row_num in range(1, train_worksheet.max_row + 1):
    row_values = tuple(train_worksheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
    nouns.append(str(row_values[0]))
    classes.append(str(row_values[1]))

# Split the data into train and test sets
train_nouns, test_nouns, train_classes, test_classes = train_test_split(nouns, classes, test_size=0.2, random_state=42, stratify=classes)

test_workbook = openpyxl.Workbook()
test_worksheet = test_workbook.active

# Write the train data to the worksheet
row_num=1
for noun, class_label in zip(train_nouns, train_classes):
    test_worksheet.cell(row=row_num, column=1, value=noun)
    test_worksheet.cell(row=row_num, column=2, value=class_label)
    row_num+=1
# Save the train workbook
test_workbook.save(output_label + '80%TrainSet.xlsx')

# Write the test data to a new Excel file
test_workbook = openpyxl.Workbook()
test_worksheet = test_workbook.active


# Write the train data to the worksheet
row_num=1
for noun, class_label in zip(test_nouns, test_classes):
    test_worksheet.cell(row=row_num, column=1, value=noun)
    test_worksheet.cell(row=row_num, column=2, value=class_label)
    row_num+=1
# Save the train workbook
test_workbook.save(output_label + '20%TestSet.xlsx')